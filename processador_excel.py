from openpyxl import load_workbook

def processar_excel(caminho_excel):
    wb = load_workbook(caminho_excel, data_only=True)
    ws = wb.active

    area = ws["B1"].value
    titulo = ws["B2"].value
    data_ref = ws["B3"].value
    introducao = ws["B4"].value

    data_ref_str = data_ref.strftime("%d/%m/%Y")
    data_segura = data_ref.strftime("%Y-%m-%d")

    cor_rgb = ws["B5"].fill.fgColor.rgb

    if cor_rgb == "FF00B050":
        cor = "verde"
    elif cor_rgb == "FFFFFF00":
        cor = "amarelo"
    elif cor_rgb == "FFFF0000":
        cor = "vermelho"
    else:
        cor = "indefinido"

    # AQUI ESTÁ A REGRA DO HORÁRIO
    # Verde continua 08-18
    # Todas as outras cores (amarelo, vermelho, etc.) agora usam 00-06
    if cor == "verde":
        periodo = f"{data_ref_str} * 08-18"
    else:
        periodo = f"{data_ref_str} * 00-06"

    impacto = "Com impacto" if cor == "vermelho" else "Sem impacto"
    riscos = impacto

    nodeList = []
    for row in ws.iter_rows(min_row=6, max_row=1000, min_col=2, max_col=2):
        cel = row[0].value
        if cel:
            nodeList.append(str(cel))
        else:
            break

    area_upper = str(area).upper()

    criado_por_map = {
        "MG": "Lúcio Flávio",
        "NE": "Alexandre Oliveira",
        "N": "Valter Júnior",
        "BASE": "Nayara Stella"
    }

    criado_por = criado_por_map.get(area_upper, "Não definido")

    contatos_vivo = {
        "BASE": ("Mario Jorge do Espírito Santo", "Gerência Acesso Móvel", "+55 71 9 9980-1340"),
        "MG": ("Paulo Ernane", "Gerência Acesso Móvel", "+55 31 98821-1589"),
        "N": ("Tássio Dantas Anaissi", "Gerência Acesso Móvel", "+55 91 99393 7766"),
        "NE": ("Leonardo Marques de Oliveira Garcia", "Gerência Acesso Móvel", "+55 85 98205 6868")
    }

    nomeVivo, respVivo, telVivo = contatos_vivo.get(area_upper, ("Contato não definido", "-", "-"))

    return {
        "area": area,
        "titulo": titulo,
        "data_ref": data_ref_str,
        "data_segura": data_segura,
        "introducao": introducao,
        "cor": cor,
        "periodo": periodo,
        "impacto": impacto,
        "riscos": riscos,
        "nodeList": nodeList,
        "criado_por": criado_por,
        "nomeVivo": nomeVivo,
        "respVivo": respVivo,
        "telVivo": telVivo
    }